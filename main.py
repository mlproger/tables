import configparser
import time
from bs4 import BeautifulSoup
import telethon
import asyncio
from telethon import events, functions, types
from telethon.tl.functions.channels import GetFullChannelRequest
from telethon.errors import SessionPasswordNeededError
from telethon.sync import TelegramClient
import pandas as pd
from pandas.io.excel import ExcelWriter
import openpyxl
from test import first_numbers, second_Numbers, email, birthsday_date

config = configparser.ConfigParser()
config.read("config.ini")

api_id = config['Telegram']['api_id']
api_hash = config['Telegram']['api_hash']

phone = config['Telegram']['phone']
username = config['Telegram']['username']

current_name = ""
        
if __name__ == '__main__':
    client = TelegramClient(username, api_id, api_hash)
    client.start()
    
    @client.on(events.NewMessage(chats=('eye_bot_auto')))
    async def normal_handler(event):
        global current_name
        if event.media != None:
            file_name = event.media.document.attributes[0].file_name
            await client.download_media(event.message, f"files/{file_name}")

            data = pd.read_excel(f"files/{file_name}", usecols=[0,1,2])
            index_c = 2
            ind = 0
            for i in data['ФИО'].tolist():
                try:
                    await client.send_message('https://t.me/GlazOfGod1337bot', f'{i}')
                    await asyncio.sleep(10)

                    if str(data['Дата рождения'].tolist()[ind]) != "NaT":
                        date = birthsday_date(current_name, str(data['ИНН'].tolist()[ind]))
                        await client.send_message('https://t.me/GlazOfGod1337bot', f'{i}, {date}')
                        await asyncio.sleep(10)
                        num2 = second_Numbers(current_name, str(data['ИНН'].tolist()[ind]))
                        email_res = email(current_name, str(data['ИНН'].tolist()[ind]))
                        wb = openpyxl.load_workbook(f"files/{file_name}")
                        ws = wb.active
                        ws[f'F{index_c}'] = email_res
                        ws[f'E{index_c}'] = num2
                        wb.save(f"files/{file_name}")

                    else:
                        date = birthsday_date(current_name, str(data['ИНН'].tolist()[ind]))
                        await client.send_message('https://t.me/GlazOfGod1337bot', f'{i}, {date}')
                        await asyncio.sleep(10)
                        num1 = first_numbers(current_name, str(data['ИНН'].tolist()[ind])),
                        wb = openpyxl.load_workbook(f"files/{file_name}")
                        num1 = first_numbers(current_name, str(data['ИНН'].tolist()[ind]))
                        ws = wb.active
                        ws[f'C{index_c}'] = date
                        ws[f'D{index_c}'] = num1
                        wb.save(f"files/{file_name}")

                    index_c+=1
                    ind += 1
                except Exception as e:
                    print(e)
                    index_c+=1
                    ind += 1
                
            await client.send_file('https://t.me/eye_bot_auto', f"files/{file_name}")

            

            

    @client.on(events.NewMessage(chats=('https://t.me/GlazOfGod1337bot')))
    async def normal_handler(event):
        global current_name
        if "**Учетная запись заблокирована до**" in event.message.text:
            await event.message.click(0)
        if event.media != None:
            name = event.media.document.attributes[0].file_name
            current_name = name
            await client.download_media(event.message, f"data/{current_name}")

        if event.message.text == "**Выберите страну, чтобы осуществить поиск**":
            await event.message.click(0)
        else:
            try:
                await event.message.click(1)
            except Exception as e:
                pass


    
    client.run_until_disconnected()